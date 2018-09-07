# -*- coding: gbk -*-

import pandas as pd
from accounts_info import *
from ret_decomposition import *
import openpyxl as ox
import utilities as ut

class output_accounts():
    def __init__(self, Asset, TotalAlpha):
        self.Asset = Asset
        self.TotalAlpha = TotalAlpha
        wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))
        sheet_accounts = wb["accounts"]
        i = 10
        while 1:
            i = i + 1
            if (sheet_accounts["A"+str(i)].value == self.Asset.StandProduct.t.yesterday.replace('-', '/')):
                break
        self.accounts_num = str(i+1)
        self.accounts_num_pre = str(i)
        wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))

    def output_sheet_accounts(self):
        wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))
        sheet_accounts = wb["accounts"]
        line_num = self.accounts_num
        line_pre = self.accounts_num_pre

        sheet_accounts["A" + line_num] = self.Asset.StandProduct.t.date.replace('-', '/')
        sheet_accounts["B" + line_num] = self.Asset.stock_holding_value_today
        sheet_accounts["C" + line_num] = self.Asset.stock_value_yesterday
        sheet_accounts["D" + line_num] = self.Asset.stock_value_today
        sheet_accounts["E" + line_num] = "=B" + line_num + "-D" + line_num
        sheet_accounts["F" + line_num] = self.Asset.money_fund
        sheet_accounts["G" + line_num] = float(self.Asset.StandProduct.csv["asset"]["Avaliable"])
        sheet_accounts["H" + line_num] = self.Asset.buyback
        sheet_accounts["I" + line_num] = float(self.Asset.StandProduct.csv["asset"]["Frozen"])
        sheet_accounts["J" + line_num] = 0
        sheet_accounts["K" + line_num] = "=" + str(self.Asset.StandProduct.csv["asset"].ix[0, "TotalAsset"]) + "+J" + line_num
        sheet_accounts["L" + line_num] = "=B" + line_num + "+G" + line_num + "+H" + line_num + "+I" + line_num + "+J" + line_num + "+F" + line_num
        sheet_accounts["M" + line_num] = float(self.Asset.StandProduct.csv["asset_fut"]["Balance"]) - self.Asset.total_closing + self.Asset.total_settle
        sheet_accounts["N" + line_num] = float(self.Asset.StandProduct.csv["asset_fut"]["CurrMargin"])
        sheet_accounts["O" + line_num] = 0
        sheet_accounts["P" + line_num] = 0
        sheet_accounts["Q" + line_num] = float(self.Asset.StandProduct.csv["asset_fut"]["Avaliable"])
        sheet_accounts["R" + line_num] = "=M" + line_num + "+O" + line_num + "+P" + line_num
        sheet_accounts["S" + line_num] = "=K" + line_num + "+R" + line_num+ "+BC" + line_num
        sheet_accounts["T" + line_num] = 0


        sheet_accounts["U" + line_num] = "=U" + line_pre + "*(AP" + line_num + "-T" + line_num + ")/AP" + line_pre
        sheet_accounts["V" + line_num] = "=S" + line_num + "-S" + line_pre
        sheet_accounts["W" + line_num] = "=V" + line_num + "/S" + line_pre
        sheet_accounts["X" + line_num] = 0
        sheet_accounts["Y" + line_num] = 0
        sheet_accounts["Z" + line_num] = 0
        sheet_accounts["AA" + line_num] = 0
        sheet_accounts["AB" + line_num] = self.Asset.total_short_value
        sheet_accounts["AC" + line_num] = self.Asset.total_long_value
        sheet_accounts["AD" + line_pre] = self.TotalAlpha.holding_beta
        sheet_accounts["AE" + line_num] = "=(AC" + line_num + "+B" + line_num + "-AB" + line_num + ")/(" + "AC" + line_num + "+B" + line_num + "+AB" + line_num + '+0.00000000000000001)'
        sheet_accounts["AF" + line_num] = "=(M" + line_num + "-N" + line_num + ")/" + "AB" + line_num
        sheet_accounts["AG" + line_num] = 0
        sheet_accounts["AH" + line_num] = 0
        sheet_accounts["AI" + line_num] = 0
        sheet_accounts["AJ" + line_num] = "=S" + line_num + "+AG" + line_num + "+AH" + line_num + "+AI" + line_num
        if self.Asset.StandProduct.product_id == 0:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+MAX((S" + line_num + "-AO" + line_pre + ")*0.2%, 50000)/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+MAX((S" + line_num + "-AO" + line_pre + ")*0.1%, 25000)/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.5%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 1:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+MAX((S" + line_num + "-AO" + line_pre + ")*0.5%, 100000)/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+MAX((S" + line_num + "-AO" + line_pre + ")*0.2%, 100000)/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.0%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 2:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.5%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.0%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 4:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.0%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+MAX(S" + line_num + "-AO" + line_pre + ")*0.05%, 25000)/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+MAX(S" + line_num + "-AO" + line_pre + ")*0.05%, 25000)/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 5:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.5%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 11:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.5%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 7:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*2.0%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.07%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.03%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 8:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.8%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 12:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.5%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = 0
        if self.Asset.StandProduct.product_id == 13:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.04%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.04%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.5%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 16:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.5%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = '=MAX(0.05%*(S' + line_num + '-AQ' + line_pre + '),50000)/365*(A' + line_num + '-A' + line_pre + ')+AL' + line_pre
        if self.Asset.StandProduct.product_id == 17:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.8%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 19:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+MAX((S" + line_num + "-AO" + line_pre + ")*0.9%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+MAX((S" + line_num + "-AO" + line_pre + ")*0.05%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.9%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 101:
            sheet_accounts["AK" + line_num] = 0
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 102:
            sheet_accounts["AK" + line_num] = 0
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = "=AM" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
        if self.Asset.StandProduct.product_id == 107:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.4%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.15%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = 0
        if self.Asset.StandProduct.product_id == 203:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.6%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*0.1%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = 0
        if self.Asset.StandProduct.product_id == 208:
            sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.7%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.2%/365*(A" + line_num + "-A" + line_pre + ")"
            sheet_accounts["AM" + line_num] = 0
        # if self.Asset.StandProduct.product_id == 301:
        #     sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+MAX((S" + line_num + "-AO" + line_pre + ")*0.05%, 100000)/365*(A" + line_num + "-A" + line_pre + ")"
        #     sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.2%/365*(A" + line_num + "-A" + line_pre + ")"
        #     sheet_accounts["AM" + line_num] = 0
        # if self.Asset.StandProduct.product_id == 309:
        #     sheet_accounts["AK" + line_num] = "=AK" + line_pre + "+MAX((S" + line_num + "-AO" + line_pre + ")*0.05%, 100000)/365*(A" + line_num + "-A" + line_pre + ")"
        #     sheet_accounts["AL" + line_num] = "=AL" + line_pre + "+(S" + line_num + "-AO" + line_pre + ")*1.2%/365*(A" + line_num + "-A" + line_pre + ")"
        #     sheet_accounts["AM" + line_num] = 0


        if self.Asset.StandProduct.product_id == 16:
            sheet_accounts["AN" + line_num] = '=MAX(0.05%*(S' + line_num + '-AQ' + line_num + '),50000)/365*(A' + line_num + '-A' + line_pre + ')+AN' + line_pre
        else:
            sheet_accounts["AN" + line_num] = "=AN" + line_pre
        sheet_accounts["AO" + line_num] = "=AK" + line_num + "+AL" + line_num + "+AM" + line_num + "+AN" + line_num
        sheet_accounts["AP" + line_num] = "=AJ" + line_num + "-AO" + line_num + "+AR" + line_num

        sheet_accounts["AR" + line_num] = "=AR" + line_pre

        sheet_accounts["AT" + line_num] = "=(U" + line_num + "-1)*365/(" + "A" + line_num + "-$A$4+1)+1"
        sheet_accounts["AU" + line_num] = "=AU" + line_pre + "*S" + line_num + "/S" + line_pre
        sheet_accounts["AV" + line_num] = "=(AU" + line_num + "-1)*365/(" + "A" + line_num + "-$A$4+1)+1"
        sheet_accounts["AW" + line_num] = "=AW" + line_pre + "*(AP" + line_num + "+AO" + line_num + "-AO" + line_pre + ")/(AP" + line_pre + ")"

        if self.Asset.StandProduct.product_id == 0:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$362)*365/(A" + line_num + "-$A$362)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$362)*365/(A" + line_num + "-$A$362)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$362"
        if self.Asset.StandProduct.product_id == 1:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$340)*365/(A" + line_num + "-$A$340)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$340)*365/(A" + line_num + "-$A$340)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$340"
        if self.Asset.StandProduct.product_id == 2:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$235)*365/(A" + line_num + "-$A$235)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$235)*365/(A" + line_num + "-$A$235)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$235"
        if self.Asset.StandProduct.product_id == 4:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-1)*365/(A" + line_num + "-DATE(2016, 9, 30))"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 5:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-1)*365/(A" + line_num + "-DATE(2016, 9, 30))"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 7:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-1)*365/(A" + line_num + "-DATE(2016, 9, 30))"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 8:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 11:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 12:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$229)*365/(A" + line_num + "-$A$229)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$229)*365/(A" + line_num + "-$A$229)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$229"
        if self.Asset.StandProduct.product_id == 13:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$229)*365/(A" + line_num + "-$A$229)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$229)*365/(A" + line_num + "-$A$229)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$229"
        if self.Asset.StandProduct.product_id == 16:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$61)*365/(A" + line_num + "-$A$61)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$61)*365/(A" + line_num + "-$A$61)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$61"
        if self.Asset.StandProduct.product_id == 17:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$64)*365/(A" + line_num + "-$A$64)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$64)*365/(A" + line_num + "-$A$64)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$64"
        if self.Asset.StandProduct.product_id == 19:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-1)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 101:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 102:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 107:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 203:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        if self.Asset.StandProduct.product_id == 208:
            sheet_accounts["AX" + line_num] = "=(U" + line_num + "-$U$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BA" + line_num] = "=(AU" + line_num + "-$AU$4)*365/(A" + line_num + "-$A$4)"
            sheet_accounts["BB" + line_num] = "=U" + line_num + "-$U$4"
        sheet_accounts["BD" + line_num] = self.Asset.holding_Hushi
        sheet_accounts["BE" + line_num] = self.Asset.holding_Shenshi
        sheet_accounts["BF" + line_num] = self.Asset.holding_sh50
        sheet_accounts["BG" + line_num] = self.Asset.holding_hs300
        sheet_accounts["BH" + line_num] = self.Asset.holding_csi500


        wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))

class output_ret_decomposition():
    def __init__(self, TotalAlpha, Alpha, Unilateral,  CrossIndex, Asset):
        self.TotalAlpha = TotalAlpha
        self.Alpha = Alpha
        self.Unilateral = Unilateral
        self.CrossIndex = CrossIndex
        self.Asset = Asset

        wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))
        sheet_ret = wb["ret_dc"]
        i = 10;
        while 1:
            i = i + 1
            if (sheet_ret["A" + str(i)].value == self.Asset.StandProduct.t.yesterday.replace('-','/')):
                break;
        self.ret_num = str(i + 1)
        self.ret_num_pre = str(i)

        sheet_accounts = wb["accounts"]
        i = 10;
        while 1:
            i = i + 1
            if (sheet_accounts["A" + str(i)].value == self.Asset.StandProduct.t.yesterday.replace('-', '/')):
                break;
        self.accounts_num = str(i + 1)
        self.accounts_num_pre = str(i)
        wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))


    def output_sheet_ret(self):
        line_num = self.ret_num
        line_num_accounts = self.accounts_num
        line_pre = self.ret_num_pre
        line_pre_accounts = self.accounts_num_pre
        wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id,self.Asset.StandProduct.t.date.replace('-', '')))
        sheet_ret = wb["ret_dc"]

        sheet_ret["A" + line_num] = self.Asset.StandProduct.t.date.replace('-', '/')
        sheet_ret["B" + line_num] = "=accounts!V" + line_num_accounts
        sheet_ret["C" + line_num] = "=D" + line_num + "+F" + line_num + "+H" + line_num + "+I" + line_num + "+O" + line_num + "+P" + line_num + "+S" + line_num + "+L" + line_num + "+J" + line_num + "+BC" + line_num
        sheet_ret["D" + line_num] = 0  #
        sheet_ret["E" + line_num] = "=" + str(self.TotalAlpha.holding_raw_ret) + "-accounts!AD" + line_pre_accounts + "*AY" + line_num
        sheet_ret["F" + line_num] = "=E" + line_num + "*accounts!D" + line_pre_accounts
        sheet_ret["G" + line_num] = "=(accounts!C" + line_num_accounts + "-accounts!D" + line_pre_accounts + ")-accounts!AD" + line_pre_accounts + "*AY" + line_num_accounts + "*accounts!D" + line_pre_accounts
        sheet_ret["H" + line_num] = self.CrossIndex.cross_index
        sheet_ret["I" + line_num] = self.TotalAlpha.premium
        sheet_ret["J" + line_num] = self.Unilateral.single_side
        sheet_ret["K" + line_num] = 0
        sheet_ret["L" + line_num] = "=accounts!F" + line_num_accounts + "-accounts!F" + line_pre_accounts + "-" + str(self.Asset.money_fund_trade_value)
        sheet_ret["M" + line_num] = self.Asset.total_buy
        sheet_ret["N" + line_num] = self.Asset.total_sell
        sheet_ret["O" + line_num] = float(self.Asset.trade_commission)
        sheet_ret["P" + line_num] = float(self.Asset.trade_fut_commission)
        sheet_ret["Q" + line_num] = self.Asset.trade_cost
        sheet_ret["R" + line_num] = self.Asset.trade_fut_cost
        sheet_ret["S" + line_num] = "=Q" + line_num + "+R" + line_num
        sheet_ret["T" + line_num] = "=B" + line_num + "-C" + line_num
        sheet_ret["Y" + line_num] = self.TotalAlpha.StandProduct.target_pre_value_dict['dvd_uncon_con50']
        sheet_ret["Z" + line_num] = self.TotalAlpha.StandProduct.target_real_ret_dict['dvd_uncon_con50']

        sheet_ret["AU" + line_num] = self.TotalAlpha.StandProduct.target_real_ret_dict['final']
        sheet_ret["AY" + line_num] = self.TotalAlpha.mix_ret


        wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))

class update():
    def __init__(self, Asset):
        self.Asset = Asset
        wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))
        sheet_accounts = wb["accounts"]
        i = 10;
        while 1:
            i = i + 1
            if (sheet_accounts["A" + str(i)].value == self.Asset.StandProduct.t.yesterday.replace('-', '/')):
                break;

        self.accounts_num = str(i + 1)
        self.accounts_num_pre = str(i)

        sheet_ret = wb["ret_dc"]
        i = 10;
        while 1:
            i = i + 1
            if (sheet_ret["A" + str(i)].value == self.Asset.StandProduct.t.yesterday.replace('-', '/')):
                break;
        self.ret_num = str(i + 1)
        self.ret_num_pre = str(i)
        wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))
    def update_summary_1(self):
        wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))
        ut.update_summary_type1(wb, self.accounts_num, self.ret_num, 10, 19, 28)
        ut.update_summary_type1(wb, str(int(self.accounts_num)-1), str(int(self.ret_num)-1), 9, 18, 27)
        ut.update_summary_type1(wb, str(int(self.accounts_num)-2), str(int(self.ret_num)-2), 8, 17, 26)
        ut.update_summary_type1(wb, str(int(self.accounts_num)-3), str(int(self.ret_num)-3), 7, 16, 25)
        ut.update_summary_type1(wb, str(int(self.accounts_num)-4), str(int(self.ret_num)-4), 6, 15, 24)
        ut.update_summary_type1(wb, str(int(self.accounts_num)-5), str(int(self.ret_num)-5), 5, 14, 23)
        wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))

    def update_summary_2(self):
        wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))
        ut.update_summary_type2(wb, self.accounts_num, self.ret_num, self.Asset.StandProduct.t.yesterday.replace('-', '/'), self.Asset.StandProduct.t.date.replace('-', '/'))
        wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.t.date.replace('-', '')))

    def update_assessreport(self):
        if self.Asset.StandProduct.product_id in [0]:
            ut.update_assessReport_0(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [2]:
            ut.update_assessReport_2(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [12]:
            ut.update_assessReport_12(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [13]:
            ut.update_assessReport_13(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [16]:
            ut.update_assessReport_16(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [17]:
            ut.update_assessReport_17(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [4]:
            ut.update_assessReport_4(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [5]:
            ut.update_assessReport_5(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [7]:
            ut.update_assessReport_7(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [11]:
            ut.update_assessReport_11(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        if self.Asset.StandProduct.product_id in [19]:
            ut.update_assessReport_19(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
        # if self.Asset.StandProduct.product_id in [8]:
        #     ut.update_assessReport_8(self.Asset.StandProduct.product_id, ur"//shiming/accounts/accounts-{}/鸣石量化对冲管理计划-{}号".format(self.Asset.StandProduct.product_id, self.Asset.StandProduct.product_id), self.Asset.StandProduct.t.date.replace('-', ''))
