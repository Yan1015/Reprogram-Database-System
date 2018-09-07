# -*- coding: gbk -*-
import pandas as pd
import os
import openpyxl as ox
import re
import xlrd
import numpy as np
import shutil

def cut_cross_fut(cross_fut, prop=None):
    columns = ['long', 'long_amount', 'short', 'short_amount', 'property', 'long_contract', 'short_contract']
    trans = pd.DataFrame({},columns=columns)

    if cross_fut in [0,"0","","0.0",0.0] or cross_fut==None:
        return trans

    cross_fut = str(cross_fut).upper().replace(' ','').replace('\n','')
    print cross_fut
    tag = 0

    for item in cross_fut.split(';'):
        l = item.split('-')
        print l
        if len(l)>2 or l==['']:
            print "cross future record type error"
            continue
        elif l[0]=='':
            m=['',0]
            n=l[1].split('*')
            tag = 1
            print m
            print n

        # trans.append(['',0,m[0][:2],float(m[1]),1])
        elif len(l)==1:
            m=l[0].split('*')
            n=['',0]
            tag = 1
        # trans.append([m[0][:2],float(m[1]),'',0,1])
        else:
            m=l[0].split('*')
            n=l[1].split('*')
            tag = 3 if m[0]==n[0] else 2
        if prop:
            tag = prop
        if tag==4:
            m[1]=float(m[1])*(-1)
        trans = trans.append(pd.DataFrame([[m[0][:2],float(m[1]),n[0][:2],float(n[1]),tag,m[0],n[0]]], columns=columns))
        print m
        print n
        print tag
        print trans

    return trans

def find_type(x):
        try:
            int(x)
        except:
            return 0
        return 1

def standard_csv_direction(uc):
    for standard_buy in [u"多",u"买","long","buy"]:
        if standard_buy in uc:
            return "buy"

    for standard_sell in [u"空",u"卖","short","sell" ,u"少"]:
        if standard_sell in uc:
            return "sell"

def check_dataframe(df): #delete all the unexpected signs
    replace_list = ['=', '"']
    for y in replace_list:
        df.columns = [x.replace(y,'') for x in df.columns]
    for x in df.columns:
        for y in replace_list:
            if df[x].dtype in (str,object):
                df[x] = df[x].str.replace(y,'')
    return df

def distinguish_commission_type(id): # 计算股票交易手续费的时候判断货币基金，回购和正常股票的差别
    id = str(id)
    pattern = re.compile(r'\b159[0-9]{1,3}')
    pattern1 = re.compile(r'\b511[0-9]{1,3}')
    l = len(pattern.findall(id)) + len(pattern1.findall(id))
    if l>0: return 0
    pattern = re.compile(r'\b150[0-9]{1,3}')
    l = len(pattern.findall(id))
    if l > 0: return 0.5
    pattern = re.compile(r'\b131[0-9]{3}')
    pattern1 = re.compile(r'\b204[0-9]{3}')
    l = len(pattern.findall(id)) + len(pattern1.findall(id))
    if l>0:return 0.7
    return 1

def extract_strategy_type(filename):
    pattern = re.compile(r'\$(.{1,})\$')
    return pattern.findall(filename)

def remove_pre_delivery_month(target_name, dm):
    target_name = str(target_name)
    pattern = re.compile(r'[A-Z]{2}([0-9]{4})\*')
    month_list = pattern.findall(target_name)
    for month in set(month_list):
        if month not in dm["delivery"].values and month in dm["pre_delivery"].values:
            pattern = re.compile('-?[A-Z]{2}'+month+'\*[0-9]{1,}')
            return pattern.sub('',target_name)
    return target_name

def set_contract_name(id):
    if id[2:4] == '00':
        return 'current'
    if id[2:4] == '01':
        return 'following'
    if id[2:4] == '02':
        return 'season'
    if id[2:4] == '03':
        return 'nextseason'

def classify_holding_type(holding_df, date):
    holding_df['str_id'] = holding_df['Id'].apply(lambda x: '0' * (6 - len(str(x))) + str(x))

    holding_df['H'] = holding_df['str_id'].apply(lambda x: 1 if x[:2:] == '60' else 0)
    holding_df['S'] = holding_df['str_id'].apply(lambda x: 1 if (x[:3:] == '000') or (x[:3:] == '002') or (x[:3:] == '300') else 0)

    df1 = pd.read_csv(r'\\SHIMING\update_file_auto\data_download\sh50\{}.csv'.format(date))
    df1['id'] = df1['coid'].apply(lambda x: int(x.split('.')[0]))
    coid_list1 = list(df1['id'])
    holding_df['sh50'] = holding_df['Id'].apply(lambda x: 1 if x in coid_list1 else 0)

    df1 = pd.read_csv(r'\\SHIMING\update_file_auto\data_download\hs300\{}.csv'.format(date))
    df1['id'] = df1['coid'].apply(lambda x: int(x.split('.')[0]))
    coid_list1 = list(df1['id'])
    holding_df['hs300'] = holding_df['Id'].apply(lambda x: 1 if x in coid_list1 else 0)

    df1 = pd.read_csv(r'\\SHIMING\update_file_auto\data_download\csi500\{}.csv'.format(date))
    df1['id'] = df1['coid'].apply(lambda x: int(x.split('.')[0]))
    coid_list1 = list(df1['id'])
    holding_df['csi500'] = holding_df['Id'].apply(lambda x: 1 if x in coid_list1 else 0)

    return holding_df

# 放在哪里都不是很妥
def download_xuntou():
    print '*****begin download xuntou*****'
    os.chdir('//shiming/accounts/CSVDownloader')
    os.system('python Downloader.py')

def download_ctp():
    print '*****begin download CTP*****'
    os.chdir('//shiming/accounts/CSVDownloader/CTPAPI_V0.95')
    os.system('python Trader.py')

def update_summary_type1(wb, sheet_accounts_num, sheet_ret_num, l1, l2, l3):
    sheet_summary = wb['summary']
    sheet_accounts = wb['accounts']
    sheet_ret = wb['ret_dc']
    l1 = str(l1)
    l2 = str(l2)
    l3 = str(l3)
    sheet_summary['A' + l1] = '=accounts!A' + sheet_accounts_num
    sheet_summary['B' + l1] = '=accounts!B' + sheet_accounts_num
    sheet_summary['C' + l1] = '=accounts!G' + sheet_accounts_num
    sheet_summary['D' + l1] = '=accounts!H' + sheet_accounts_num
    sheet_summary['E' + l1] = '=accounts!I' + sheet_accounts_num
    sheet_summary['F' + l1] = '=accounts!J' + sheet_accounts_num
    sheet_summary['G' + l1] = '=accounts!K' + sheet_accounts_num
    sheet_summary['H' + l1] = '=accounts!F' + sheet_accounts_num
    sheet_summary['I' + l1] = '=accounts!M' + sheet_accounts_num
    sheet_summary['J' + l1] = '=accounts!N' + sheet_accounts_num
    sheet_summary['K' + l1] = '=accounts!O' + sheet_accounts_num
    sheet_summary['L' + l1] = '=accounts!P' + sheet_accounts_num
    sheet_summary['M' + l1] = '=accounts!Q' + sheet_accounts_num
    sheet_summary['N' + l1] = '=accounts!R' + sheet_accounts_num
    sheet_summary['O' + l1] = '=accounts!S' + sheet_accounts_num
    sheet_summary['P' + l1] = '=accounts!U' + sheet_accounts_num
    sheet_summary['Q' + l1] = '=accounts!AT' + sheet_accounts_num
    sheet_summary['R' + l1] = '=accounts!AS' + sheet_accounts_num
    sheet_summary['S' + l1] = '=accounts!AU' + sheet_accounts_num
    sheet_summary['T' + l1] = '=accounts!AX' + sheet_accounts_num
    sheet_summary['U' + l1] = '=accounts!BA' + sheet_accounts_num
    sheet_summary['V' + l1] = '=accounts!BB' + sheet_accounts_num
    sheet_summary['W' + l1] = '=ret_dc!B' + sheet_ret_num
    sheet_summary['X' + l1] = '=accounts!W' + sheet_accounts_num
    sheet_summary['Y' + l1] = '=accounts!X' + sheet_accounts_num
    sheet_summary['Z' + l1] = '=accounts!Y' + sheet_accounts_num
    sheet_summary['AA' + l1] = '=accounts!Z' + sheet_accounts_num
    sheet_summary['AB' + l1] = '=accounts!AA' + sheet_accounts_num
    sheet_summary['AC' + l1] = '=accounts!AD' + sheet_accounts_num

    sheet_summary['A' + l2] = '=ret_dc!A' + sheet_ret_num
    sheet_summary['B' + l2] = '=ret_dc!U' + sheet_ret_num
    sheet_summary['C' + l2] = '=ret_dc!V' + sheet_ret_num
    sheet_summary['D' + l2] = '=ret_dc!W' + sheet_ret_num
    sheet_summary['E' + l2] = '=ret_dc!X' + sheet_ret_num
    sheet_summary['F' + l2] = '=ret_dc!Y' + sheet_ret_num
    sheet_summary['G' + l2] = '=ret_dc!Z' + sheet_ret_num
    sheet_summary['H' + l2] = '=ret_dc!AA' + sheet_ret_num
    sheet_summary['I' + l2] = '=ret_dc!AB' + sheet_ret_num
    sheet_summary['J' + l2] = '=ret_dc!AC' + sheet_ret_num
    sheet_summary['K' + l2] = '=ret_dc!AD' + sheet_ret_num
    sheet_summary['L' + l2] = '=ret_dc!AE' + sheet_ret_num
    sheet_summary['M' + l2] = '=ret_dc!AF' + sheet_ret_num
    sheet_summary['N' + l2] = '=ret_dc!AG' + sheet_ret_num
    sheet_summary['O' + l2] = '=ret_dc!AH' + sheet_ret_num
    sheet_summary['P' + l2] = '=ret_dc!AI' + sheet_ret_num
    sheet_summary['Q' + l2] = '=ret_dc!AJ' + sheet_ret_num
    sheet_summary['R' + l2] = '=ret_dc!AK' + sheet_ret_num
    sheet_summary['S' + l2] = '=ret_dc!AL' + sheet_ret_num
    sheet_summary['T' + l2] = '=ret_dc!AM' + sheet_ret_num
    sheet_summary['U' + l2] = '=ret_dc!AN' + sheet_ret_num
    sheet_summary['V' + l2] = '=ret_dc!AO' + sheet_ret_num
    sheet_summary['W' + l2] = '=ret_dc!AP' + sheet_ret_num
    sheet_summary['X' + l2] = '=ret_dc!AQ' + sheet_ret_num
    sheet_summary['Y' + l2] = '=ret_dc!AR' + sheet_ret_num
    sheet_summary['Z' + l2] = '=ret_dc!AS' + sheet_ret_num
    sheet_summary['AA' + l2] = '=ret_dc!AT' + sheet_ret_num
    sheet_summary['AB' + l2] = '=ret_dc!AU' + sheet_ret_num
    sheet_summary['AC' + l2] = '=accounts!BD' + sheet_accounts_num
    sheet_summary['AD' + l2] = '=accounts!BE' + sheet_accounts_num
    sheet_summary['AE' + l2] = '=accounts!BF' + sheet_accounts_num
    sheet_summary['AF' + l2] = '=accounts!BG' + sheet_accounts_num
    sheet_summary['AG' + l2] = '=accounts!BH' + sheet_accounts_num

    sheet_summary['A' + l3] = '=ret_dc!A' + sheet_ret_num
    sheet_summary['B' + l3] = '=ret_dc!B' + sheet_ret_num
    sheet_summary['C' + l3] = '=ret_dc!C' + sheet_ret_num
    sheet_summary['D' + l3] = '=ret_dc!D' + sheet_ret_num
    sheet_summary['E' + l3] = '=ret_dc!E' + sheet_ret_num
    sheet_summary['F' + l3] = '=ret_dc!F' + sheet_ret_num
    sheet_summary['G' + l3] = '=ret_dc!H' + sheet_ret_num
    sheet_summary['H' + l3] = '=ret_dc!I' + sheet_ret_num
    sheet_summary['I' + l3] = '=ret_dc!O' + sheet_ret_num + '+ret_dc!P' + sheet_ret_num
    sheet_summary['J' + l3] = '=ret_dc!Q' + sheet_ret_num
    sheet_summary['K' + l3] = '=ret_dc!R' + sheet_ret_num
    sheet_summary['L' + l3] = '=ret_dc!S' + sheet_ret_num
    sheet_summary['M' + l3] = '=ret_dc!T' + sheet_ret_num



def update_summary_type2(wb, sheet_accounts_num, sheet_ret_num, yesterday, date):
    sheet_summary = wb['summary']
    sheet_accounts = wb['accounts']
    sheet_ret = wb['ret_dc']
    i = 10
    while 1:
        i = i + 1
        if (sheet_summary["A" + str(i)].value == yesterday):
            break
    l = str(i + 1)
    sheet_summary['A' + l] = date
    sheet_summary['G' + l] = '=accounts!K' + sheet_accounts_num
    sheet_summary['M' + l] = '=accounts!R' + sheet_accounts_num
    sheet_summary['N' + l] = '=accounts!S' + sheet_accounts_num
    sheet_summary['O' + l] = '=accounts!U' + sheet_accounts_num
    sheet_summary['P' + l] = '=accounts!AU' + sheet_accounts_num
    sheet_summary['Q' + l] = '=accounts!AT' + sheet_accounts_num
    sheet_summary['R' + l] = '=accounts!AZ' + sheet_accounts_num
    sheet_summary['AB' + l] = '=accounts!AX' + sheet_accounts_num
    sheet_summary['AC' + l] = '=accounts!BA' + sheet_accounts_num
    sheet_summary['AD' + l] = '=accounts!BB' + sheet_accounts_num

def update_yangyang_index(date):
    root_path = r"//shiming/accounts/python/common_data"
    file = r'future_wind.csv'
    future_df = pd.read_csv(r'{}/{}'.format(root_path, file))
    future = pd.read_csv(r'{}/{}'.format(root_path, file))

    future_df = future_df[future_df['date'] == date]
    df = pd.read_csv(r'{}/{}'.format('//shiming/accounts/Dividend Prediction', 'wind_future.csv'))
    dd = df

    columns = ['date', 'type1', 'target1', 'type2', 'target2', 'type3', 'target3']
    df1 = pd.DataFrame(np.random.randn(1, 7), index=[0], columns=columns)
    d1 = future_df[future_df['type'] == 'IH']

    e1 = d1.iloc[0, 5]
    d2 = future_df[future_df['type'] == 'IC']
    e2 = d2.iloc[0, 5]
    d3 = future_df[future_df['type'] == 'IF']
    e3 = d3.iloc[0, 5]
    r = d1.iloc[0, 7]

    df1.iloc[0, 0] = r
    df1.iloc[0, 1] = 'IH'
    df1.iloc[0, 2] = e1
    df1.iloc[0, 3] = 'IC'
    df1.iloc[0, 4] = e2
    df1.iloc[0, 5] = 'IF'
    df1.iloc[0, 6] = e3
    df = pd.concat([df, df1])
    if date not in dd['date'].values:
        df.to_csv(r'{}/{}'.format('//shiming/accounts/Dividend Prediction', 'wind_future.csv'), encoding='gbk', index=False)

def copy_to_account(id, date):
    accounts_path = r'\\SHIMING\accounts\accounts-{}\{}'.format(id, date)
    python_path = r'\\SHIMING\accounts\python\YZJ\accounts\accounts-{}\{}'.format(id, date)
    try:
        shutil.rmtree(accounts_path)
    except:
        pass
    if not os.path.exists(accounts_path):
        shutil.copytree(python_path, accounts_path)

def update_assessReport_0(product_id, assessreport_path, date):
    folder_list = list(os.walk(assessreport_path))[0][1]
    f_list = []
    d_list = []
    for item in folder_list:
        pattern = re.compile(ur'鸣石.*')
        if len(pattern.findall(item)) != 0:
            if len(item.split(u'化 ')) > 1:
                if item.split(u'化 ')[1] == '0309':
                    continue
                f_list = f_list + [item]
                if item.split(u'化 ')[1][0:4] != '2016':
                    d_list = d_list + [int('2016' + item.split(u'化 ')[1])]
                else:
                    d_list = d_list + [int(item.split(u'化 ')[1])]
            if len(item.split(u'石0')) > 1:
                f_list = f_list + [item]
                if item.split(u'石')[1][0:4] != '2016':
                    d_list = d_list + [int('2016' + item.split(u'石')[1])]
                else:
                    d_list = d_list + [int(item.split(u'石')[1])]

    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)
    print date_dict

    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if (key > last_date) and (key <= int(date.replace('/',''))):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}/3C0085委托资产资产估值表.xls'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name('Sheet1')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'上海活期存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'上海活期存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'清算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
            if u'存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['E' + l] = 0
            index = sheet1.col_values(0).index(u'证券投资合计')
            sheet_assessReport['F' + l] = sheet1.cell(index, 7).value

            if u'其中股票投资' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'其中股票投资')
                sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['G' + l] = 0
            index = sheet1.col_values(0).index(u'其中基金投资')
            sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
            if u'投资顾问费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'投资顾问费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['N' + l] = 0

            if u'应收利息' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应收利息')
                sheet_assessReport['AS' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['AS' + l] = 0

            if u'应付赎回款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付赎回款')
                sheet_assessReport['AT' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['AT' + l] = 0

            index = sheet1.col_values(0).index(u'今日单位净值：')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'资产资产净值：' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'资产资产净值：')
                sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l


            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_2(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    dir_list = list(os.walk(assessreport_path))[0][1]
    f_list = []
    for item in file_list:
        pattern = re.compile(ur'证券投资基金估值表_鸣石投资量化2期证券投资基金_.*')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
    d_list = []
    for item in f_list:
        d = item.split('_')[-1].split('.')[0]
        d_list = d_list + [int(d.replace('-',''))]
    for item in dir_list:
        pattern = re.compile(ur'鸣石估值表')
        if len(pattern.findall(item)) != 0:
            name = list(os.walk(u'{}/{}'.format(assessreport_path, item)))[0][2][0]
            if int(name.split('.')[0].split('_')[-1].replace('-', '')) not in d_list:
                f_list = f_list + [u'{}/{}'.format(item, name)]
                d_list = d_list + [int(name.split('.')[0].split('_')[-1].replace('-', ''))]

    print f_list
    print d_list

    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)


    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if (key > last_date) and (key <= int(date.replace('/', ''))):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name('Sheet1')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'期货清算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
            if u'存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['E' + l] = 0
            index = sheet1.col_values(0).index(u'证券投资合计:')
            sheet_assessReport['F' + l] = sheet1.cell(index, 7).value

            index = sheet1.col_values(0).index(u'其中股票投资:')
            sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
            index = sheet1.col_values(0).index(u'其中基金投资:')
            sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金账户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
            if u'应付投资顾问费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付投资顾问费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(0).index(u'今日单位净值：')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'基金资产净值:' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'基金资产净值:')
                sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))


# def update_assessReport_7(product_id, assessreport_path, date):
#     file_list = list(os.walk(assessreport_path))[0][2]
#     dir_list = list(os.walk(assessreport_path))[0][1]
#     f_list = []
#     k_list = []
#     for item in file_list:
#         pattern = re.compile(ur'.*.xls')
#         if len(pattern.findall(item)) != 0:
#             f_list = f_list + [item]
#             k_list = k_list + [item]
#     for item in dir_list:
#         pattern = re.compile(ur'鸣石[0-9]{4}')
#         if len(pattern.findall(item)) != 0:
#             ff_list = list(os.walk(ur'{}/{}'.format(assessreport_path, item)))[0][2]
#             for ite in ff_list:
#                 pattern = re.compile(ur'证券投资基金估值表_银河资本鸣石量化7号_.*')
#                 if len(pattern.findall(ite)) != 0:
#                     f_list = f_list + [ite]
#                     k_list = k_list + [ur'{}/{}'.format(item, ite)]
#     d_list = []
#     print f_list
#
#     for item in f_list:
#         d = item.split('_')[-1].split('.')[0]
#         d_list = d_list + [int(d.replace('-', ''))]
#     print d_list
#     date_dict = dict(zip(d_list, k_list))
#     date_dict = sorted(date_dict.iteritems(), key=lambda x: x[0], reverse=False)
#     print date_dict
#
#     wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
#     sheet_assessReport = wb["assessReport"]
#     sheet_accounts = wb['accounts']
#     i = 3
#     while 1:
#         i = i + 1
#         if not sheet_assessReport["A" + str(i)].value:
#             break
#     l = str(i)
#     last_date = sheet_assessReport["A" + str(i - 1)].value
#     last_date = last_date.replace('/', '')
#     last_date = int(last_date)
#     for key, item in date_dict:
#         print key
#         print item
#         if (key > last_date) and (key <= int(date.replace('/', ''))):
#             assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
#             sheet1 = assessReport_wb.sheet_by_name('Sheet1')
#             DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
#             sheet_assessReport['A' + l] = DATE
#             if u'活期存款' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'活期存款')
#                 sheet_assessReport['B' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['B' + l] = 0
#             sheet_assessReport['C' + l] = '=D' + l + '+E' + l
#             index = sheet1.col_values(1).index(u'清算备付金')
#             sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
#             if u'存出保证金' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'存出保证金')
#                 sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['E' + l] = 0
#             index = sheet1.col_values(0).index(u'证券投资合计:')
#             sheet_assessReport['F' + l] = sheet1.cell(index, 7).value
#
#             index = sheet1.col_values(0).index(u'其中股票投资:')
#             sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
#             index = sheet1.col_values(0).index(u'其中基金投资:')
#             sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
#             if u'买入返售金额资产' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'买入返售金额资产')
#                 sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['I' + l] = 0
#
#             if u'三方存款账户' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'三方存款账户')
#                 sheet_assessReport['J' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['J' + l] = 0
#
#             if u'理财产品' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'理财产品')
#                 sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['K' + l] = 0
#
#             index = sheet1.col_values(1).index(u'应付托管费')
#             sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
#             if u'应付投资顾问费' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'应付投资顾问费')
#                 sheet_assessReport['M' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['M' + l] = 0
#             if u'应付管理人报酬' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'应付管理人报酬')
#                 sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['N' + l] = 0
#             index = sheet1.col_values(0).index(u'基金单位净值：')
#             sheet_assessReport['O' + l] = sheet1.cell(index, 1).value
#
#             if u'基金资产净值:' in sheet1.col_values(0):
#                 index = sheet1.col_values(0).index(u'基金资产净值:')
#                 sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['P' + l] = 0
#
#             index = sheet1.col_values(0).index(u'实收资本')
#             sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value
#
#             print DATE
#             i = 2;
#             while 1:
#                 i = i + 1
#                 if (sheet_accounts["A" + str(i)].value == DATE):
#                     break;
#
#             accounts_num = str(i)
#             print accounts_num
#
#             sheet_assessReport['R' + l] = '=accounts!M' + accounts_num
#
#             sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
#             sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
#             sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
#             sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
#             sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
#             sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
#             sheet_assessReport['Y' + l] = 0
#             sheet_assessReport['Z' + l] = 0
#             sheet_assessReport['AA' + l] = 0
#             sheet_assessReport['AB' + l] = 0
#             sheet_assessReport['AC' + l] = 0
#             sheet_assessReport['AD' + l] = 0
#             sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
#             sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
#             sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num
#
#             sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num
#
#             sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
#             sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
#             sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
#             sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
#             sheet_assessReport['AM' + l] = 0
#             sheet_assessReport['AN' + l] = 0
#             sheet_assessReport['AO' + l] = 0
#             sheet_assessReport['AP' + l] = 0
#             sheet_assessReport['AQ' + l] = '=AL' + l + '+AM' + l
#             if key >= 20160628:
#                 input_path = "{}/{}/{}".format("//shiming/accounts/CSVDownloader/Download/XT7", str(key)[4:8], 'asset_l.csv')
#                 asset = pd.read_csv(input_path, encoding='gbk')  # , usecols=asset_columns
#                 print asset
#                 sheet_assessReport['Z' + l] = asset.loc[0, u'总资产']
#
#                 trade_day_list = list(pd.read_csv('//SHIMING/accounts/python/trading_day_list.csv', names=None).iloc[:, 0])
#                 ddate = str(key)[:4] + '-' + str(key)[4:6] + '-' + str(key)[6:8]
#                 next_date = trade_day_list[trade_day_list.index(ddate) + 1]
#                 input_file = 'asset_fut_l.csv'
#                 input_path = "{}/{}/{}".format("//shiming/accounts/CSVDownloader/Download/XT7", next_date.replace('-', '')[4:8], input_file)
#                 df = pd.read_csv(input_path,encoding='gbk')
#                 print df
#                 sheet_assessReport['Y' + l] = df.loc[0, u'期初权益']
#             last_date = key
#             l = str(int(l) + 1)
#     wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
#
# def update_assessReport_11(product_id, assessreport_path, date):
#     file_list = list(os.walk(assessreport_path))[0][2]
#     dir_list = list(os.walk(assessreport_path))[0][1]
#     f_list = []
#     k_list = []
#     for item in file_list:
#         pattern = re.compile(ur'.*.xls')
#         if len(pattern.findall(item)) != 0:
#             f_list = f_list + [item]
#             k_list = k_list + [item]
#     for item in dir_list:
#         pattern = re.compile(ur'鸣石[0-9]{4}')
#         if len(pattern.findall(item)) != 0:
#             ff_list = list(os.walk(ur'{}/{}'.format(assessreport_path, item)))[0][2]
#             for ite in ff_list:
#                 pattern = re.compile(ur'证券投资基金估值表_银河资本鸣石量化11号_.*')
#                 if len(pattern.findall(ite)) != 0:
#                     f_list = f_list + [ite]
#                     k_list = k_list + [ur'{}/{}'.format(item, ite)]
#     d_list = []
#     print f_list
#
#
#     for item in f_list:
#         d = item.split('_')[-1].split('.')[0]
#         d_list = d_list + [int(d.replace('-',''))]
#     print d_list
#     date_dict = dict(zip(d_list, k_list))
#     date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)
#     print date_dict
#
#     wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
#     sheet_assessReport = wb["assessReport"]
#     sheet_accounts = wb['accounts']
#     i = 3
#     while 1:
#         i = i + 1
#         if not sheet_assessReport["A" + str(i)].value:
#             break
#     l = str(i)
#     last_date = sheet_assessReport["A" + str(i - 1)].value
#     last_date = last_date.replace('/', '')
#     last_date = int(last_date)
#     for key, item in date_dict:
#         print key
#         print item
#         if (key > last_date) and (key <= int(date.replace('/',''))):
#             assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
#             sheet1 = assessReport_wb.sheet_by_name('Sheet1')
#             DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
#             sheet_assessReport['A' + l] = DATE
#             if u'活期存款' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'活期存款')
#                 sheet_assessReport['B' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['B' + l] = 0
#             sheet_assessReport['C' + l] = '=D' + l + '+E' + l
#             index = sheet1.col_values(1).index(u'清算备付金')
#             sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
#             if u'存出保证金' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'存出保证金')
#                 sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['E' + l] = 0
#             index = sheet1.col_values(0).index(u'证券投资合计:')
#             sheet_assessReport['F' + l] = sheet1.cell(index, 7).value
#
#             index = sheet1.col_values(0).index(u'其中股票投资:')
#             sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
#             index = sheet1.col_values(0).index(u'其中基金投资:')
#             sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
#             if u'买入返售金额资产' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'买入返售金额资产')
#                 sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['I' + l] = 0
#
#             if u'三方存款账户' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'三方存款账户')
#                 sheet_assessReport['J' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['J' + l] = 0
#
#             if u'理财产品' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'理财产品')
#                 sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['K' + l] = 0
#
#             index = sheet1.col_values(1).index(u'应付托管费')
#             sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
#             if u'应付投资顾问费' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'应付投资顾问费')
#                 sheet_assessReport['M' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['M' + l] = 0
#             if u'应付管理人报酬' in sheet1.col_values(1):
#                 index = sheet1.col_values(1).index(u'应付管理人报酬')
#                 sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['N' + l] = 0
#             index = sheet1.col_values(0).index(u'基金单位净值：')
#             sheet_assessReport['O' + l] = sheet1.cell(index, 1).value
#
#             if u'基金资产净值:' in sheet1.col_values(0):
#                 index = sheet1.col_values(0).index(u'基金资产净值:')
#                 sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
#             else:
#                 sheet_assessReport['P' + l] = 0
#
#             index = sheet1.col_values(0).index(u'实收资本')
#             sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value
#
#             print DATE
#             i = 2;
#             while 1:
#                 i = i + 1
#                 if (sheet_accounts["A" + str(i)].value == DATE):
#                     break;
#
#             accounts_num = str(i)
#             print accounts_num
#
#             sheet_assessReport['R' + l] = '=accounts!M' + accounts_num
#
#             sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
#             sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
#             sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
#             sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
#             sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
#             sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
#             sheet_assessReport['Y' + l] = 0
#             sheet_assessReport['Z' + l] = 0
#             sheet_assessReport['AA' + l] = 0
#             sheet_assessReport['AB' + l] = 0
#             sheet_assessReport['AC' + l] = 0
#             sheet_assessReport['AD' + l] = 0
#             sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
#             sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
#             sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num
#
#             sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num
#
#             sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
#             sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
#             sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
#             sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
#             sheet_assessReport['AM' + l] = 0
#             sheet_assessReport['AN' + l] = 0
#             sheet_assessReport['AO' + l] = 0
#             sheet_assessReport['AP' + l] = 0
#             sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l
#             if key >= 20160627:
#                 if key in [20160805, 20160808, 20160815]:
#                     continue
#                 if key >= 20160818:
#                     input_path = "{}/{}/{}".format("//shiming/accounts/CSVDownloader/Download/XT11", str(key), 'holding.xls')
#                 else:
#                     input_path = "{}/{}/{}".format("//shiming/accounts/CSVDownloader/Download/XT11", str(key)[4:8], 'holding.xls')
#                 asset = pd.read_csv(input_path, sep='\t', encoding='gbk', nrows=1)  # , usecols=asset_columns
#                 a = asset.ix[0, 5]
#                 if isinstance(a, (float)):
#                     sheet_assessReport['Z' + l] = asset.ix[0, 5]
#                 else:
#                     sheet_assessReport['Z' + l] = asset.ix[0, 5].replace('=', '').replace(r'"', '')
#
#                 trade_day_list = list(pd.read_csv('//SHIMING/accounts/python/trading_day_list.csv', names=None).iloc[:, 0])
#                 ddate = str(key)[:4] + '-' + str(key)[4:6] + '-' + str(key)[6:8]
#                 next_date = trade_day_list[trade_day_list.index(ddate) + 1]
#                 input_file = 'asset_fut_l.txt'
#                 if int(next_date.replace('-','')) >= 20160818:
#                     input_path = "{}/{}/{}".format("//shiming/accounts/CSVDownloader/Download/XT11", next_date.replace('-', ''), input_file)
#                 else:
#                     input_path = "{}/{}/{}".format("//shiming/accounts/CSVDownloader/Download/XT11", next_date.replace('-','')[4:8], input_file)
#                 f = open(input_path, 'r')
#                 lines = f.readlines()
#                 a = float(lines[8].strip().split(' ')[-1].replace(',', ''))
#                 sheet_assessReport['Y' + l] = a
#             last_date = key
#             l = str(int(l) + 1)
#     wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_12(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    dir_list = list(os.walk(assessreport_path))[0][1]
    f_list = []
    k_list = []
    for item in file_list:
        pattern = re.compile(ur'.*.xls')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
            k_list = k_list + [item]
    for item in dir_list:
        pattern = re.compile(ur'鸣石[0-9]{4}')
        if len(pattern.findall(item)) != 0:
            ff_list = list(os.walk(ur'{}/{}'.format(assessreport_path, item)))[0][2]
            for ite in ff_list:
                pattern = re.compile(ur'证券投资基金估值表_鸣石量化投资12号证券投资基金_.*')
                if len(pattern.findall(ite)) != 0:
                    f_list = f_list + [ite]
                    k_list = k_list + [ur'{}/{}'.format(item, ite)]
    d_list = []
    print f_list

    for item in f_list:
        d = item.split('_')[-1].split('.')[0]
        d_list = d_list + [int(d.replace('-', ''))]
    print d_list
    date_dict = dict(zip(d_list, k_list))
    date_dict = sorted(date_dict.iteritems(), key=lambda x: x[0], reverse=False)
    print date_dict

    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if (key > last_date) and (key <= int(date.replace('/', ''))):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name('Sheet1')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'期货清算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
            if u'期货交易存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'期货交易存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['E' + l] = 0
            index = sheet1.col_values(0).index(u'证券投资合计:')
            sheet_assessReport['F' + l] = sheet1.cell(index, 7).value

            index = sheet1.col_values(0).index(u'其中股票投资:')
            sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
            index = sheet1.col_values(0).index(u'其中基金投资:')
            sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金账户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
            if u'应付投资顾问费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付投资顾问费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(0).index(u'今日单位净值：')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'基金资产净值:' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'基金资产净值:')
                sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_13(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    dir_list = list(os.walk(assessreport_path))[0][1]
    f_list = []
    k_list = []
    for item in file_list:
        pattern = re.compile(ur'SE8101_平安道远ＳＴＡＲ２号.*.xls')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
            k_list = k_list + [item]
    for item in dir_list:
        pattern = re.compile(ur'鸣石[0-9]{4}')
        if len(pattern.findall(item)) != 0:
            ff_list = list(os.walk(ur'{}/{}'.format(assessreport_path, item)))[0][2]
            for ite in ff_list:
                pattern = re.compile(ur'SE8101_平安道远ＳＴＡＲ２号.*')
                if len(pattern.findall(ite)) != 0:
                    f_list = f_list + [ite]
                    k_list = k_list + [ur'{}/{}'.format(item, ite)]
    d_list = []
    print f_list

    for item in f_list:
        d = item.split(u'号')[-1].split('.')[0]
        d_list = d_list + [int(d)]
    print d_list
    date_dict = dict(zip(d_list, k_list))
    date_dict = sorted(date_dict.iteritems(), key=lambda x: x[0], reverse=False)
    print date_dict

    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if (key > 20160304) and (key < 20160315):
            continue
        if (key > last_date) and (key <= int(date.replace('/', ''))):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name('Sheet1')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(2):
                index = sheet1.col_values(2).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 8).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(2).index(u'清算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 8).value
            if u'期货存出保证金' in sheet1.col_values(2):
                index = sheet1.col_values(2).index(u'期货存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 8).value
            else:
                sheet_assessReport['E' + l] = 0
            index = sheet1.col_values(1).index(u'证券投资合计:')
            sheet_assessReport['F' + l] = sheet1.cell(index, 8).value

            index = sheet1.col_values(1).index(u'其中股票投资:')
            sheet_assessReport['G' + l] = sheet1.cell(index, 8).value
            index = sheet1.col_values(1).index(u'其中基金投资:')
            sheet_assessReport['H' + l] = sheet1.cell(index, 8).value
            if u'买入返售金额资产' in sheet1.col_values(2):
                index = sheet1.col_values(2).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 8).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'券商保证金账户' in sheet1.col_values(2):
                index = sheet1.col_values(2).index(u'券商保证金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 8).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(2):
                index = sheet1.col_values(2).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 8).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(2).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 8).value
            if u'应付固定投资顾问费' in sheet1.col_values(2):
                index = sheet1.col_values(2).index(u'应付固定投资顾问费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 8).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付基金运营服务费' in sheet1.col_values(2):
                index = sheet1.col_values(2).index(u'应付基金运营服务费')
                sheet_assessReport['N' + l] = sheet1.cell(index, 8).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(1).index(u'今日单位净值：')
            sheet_assessReport['O' + l] = sheet1.cell(index, 2).value

            if u'基金资产净值:' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'基金资产净值:')
                sheet_assessReport['P' + l] = sheet1.cell(index, 8).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(1).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 8).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_16(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    f_list = []
    for item in file_list:
        pattern = re.compile(ur'证券投资基金估值表_鸣石量化多策略二号私募证券投资基金_.*')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
    d_list = []
    print f_list
    for item in f_list:
        d = item.split('_')[-1].split('.')[0]
        d_list = d_list + [int(d.replace('-',''))]
    print d_list
    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)


    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if key in [20160731,20161228]:
            continue
        if (key > last_date) and (key <= int(date.replace('/', ''))):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name('Sheet1')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'期货清算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
            if u'存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['E' + l] = 0
            index = sheet1.col_values(0).index(u'证券投资合计:')
            sheet_assessReport['F' + l] = sheet1.cell(index, 7).value

            index = sheet1.col_values(0).index(u'其中股票投资:')
            sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
            index = sheet1.col_values(0).index(u'其中基金投资:')
            sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金账户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
            if u'应付估值服务费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付估值服务费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(0).index(u'今日单位净值：')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'基金资产净值:' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'基金资产净值:')
                sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_17(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    f_list = []
    for item in file_list:
        pattern = re.compile(ur'SH4721_巨禄鸣石稳健17号证券投资基金.*')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
    d_list = []
    print f_list
    for item in f_list:
        d = item.split(u'金')[1].split('.')[0]
        d_list = d_list + [int(d)]
    print d_list
    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)


    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    print last_date
    for key, item in date_dict:
        print key
        print item
        if key < 20161216:
            if (key > last_date) and (key <= int(date.replace('/', ''))):
                assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
                sheet1 = assessReport_wb.sheet_by_name('Sheet1')
                DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
                sheet_assessReport['A' + l] = DATE
                if u'银行存款' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'银行存款')
                    sheet_assessReport['B' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['B' + l] = 0
                sheet_assessReport['C' + l] = '=D' + l + '+E' + l
                index = sheet1.col_values(2).index(u'清算备付金')
                sheet_assessReport['D' + l] = sheet1.cell(index, 8).value
                if u'期货存出保证金' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'期货存出保证金')
                    sheet_assessReport['E' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['E' + l] = 0
                index = sheet1.col_values(1).index(u'证券投资合计:')
                sheet_assessReport['F' + l] = sheet1.cell(index, 8).value

                index = sheet1.col_values(1).index(u'其中股票投资:')
                sheet_assessReport['G' + l] = sheet1.cell(index, 8).value
                index = sheet1.col_values(1).index(u'其中基金投资:')
                sheet_assessReport['H' + l] = sheet1.cell(index, 8).value
                if u'买入返售金额资产' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'买入返售金额资产')
                    sheet_assessReport['I' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['I' + l] = 0

                if u'券商保证金账户' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'券商保证金账户')
                    sheet_assessReport['J' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['J' + l] = 0

                if u'其它投资' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'其它投资')
                    sheet_assessReport['K' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['K' + l] = 0

                index = sheet1.col_values(2).index(u'应付托管费')
                sheet_assessReport['L' + l] = sheet1.cell(index, 8).value
                if u'应付基金运营服务费' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'应付基金运营服务费')
                    sheet_assessReport['M' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['M' + l] = 0
                if u'应付管理人报酬' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'应付管理人报酬')
                    sheet_assessReport['N' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['N' + l] = 0
                index = sheet1.col_values(1).index(u'今日单位净值：')
                sheet_assessReport['O' + l] = sheet1.cell(index, 2).value

                if u'基金资产净值:' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'基金资产净值:')
                    sheet_assessReport['P' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['P' + l] = 0

                index = sheet1.col_values(1).index(u'实收资本')
                sheet_assessReport['Q' + l] = sheet1.cell(index, 8).value

                print DATE
                i = 2;
                while 1:
                    i = i + 1
                    if (sheet_accounts["A" + str(i)].value == DATE):
                        break;

                accounts_num = str(i)
                print accounts_num

                sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

                sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
                sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
                sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
                sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
                sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
                sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
                sheet_assessReport['Y' + l] = 0
                sheet_assessReport['Z' + l] = 0
                sheet_assessReport['AA' + l] = 0
                sheet_assessReport['AB' + l] = 0
                sheet_assessReport['AC' + l] = 0
                sheet_assessReport['AD' + l] = 0
                sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
                sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
                sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

                sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

                sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
                sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
                sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
                sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
                sheet_assessReport['AM' + l] = 0
                sheet_assessReport['AN' + l] = 0
                sheet_assessReport['AO' + l] = 0
                sheet_assessReport['AP' + l] = 0
                sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

                last_date = key
                l = str(int(l) + 1)
        else:
            if (key > last_date) and (key <= int(date.replace('/', ''))):
                assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
                sheet1 = assessReport_wb.sheet_by_name('Sheet1')
                DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
                sheet_assessReport['A' + l] = DATE
                if u'银行存款' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'银行存款')
                    sheet_assessReport['B' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['B' + l] = 0
                sheet_assessReport['C' + l] = '=D' + l + '+E' + l
                index = sheet1.col_values(1).index(u'清算备付金')
                sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
                if u'期货存出保证金' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'期货存出保证金')
                    sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['E' + l] = 0
                index = sheet1.col_values(0).index(u'证券投资合计:')
                sheet_assessReport['F' + l] = sheet1.cell(index, 7).value

                index = sheet1.col_values(0).index(u'其中股票投资:')
                sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
                index = sheet1.col_values(0).index(u'其中基金投资:')
                sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
                if u'买入返售金额资产' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'买入返售金额资产')
                    sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['I' + l] = 0

                if u'券商保证金账户' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'券商保证金账户')
                    sheet_assessReport['J' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['J' + l] = 0

                if u'其它投资' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'其它投资')
                    sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['K' + l] = 0

                index = sheet1.col_values(1).index(u'应付托管费')
                sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
                if u'应付基金运营服务费' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'应付基金运营服务费')
                    sheet_assessReport['M' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['M' + l] = 0
                if u'应付管理人报酬' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'应付管理人报酬')
                    sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['N' + l] = 0
                index = sheet1.col_values(0).index(u'今日单位净值：')
                sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

                if u'基金资产净值:' in sheet1.col_values(0):
                    index = sheet1.col_values(0).index(u'基金资产净值:')
                    sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['P' + l] = 0

                index = sheet1.col_values(0).index(u'实收资本')
                sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value

                print DATE
                i = 2;
                while 1:
                    i = i + 1
                    if (sheet_accounts["A" + str(i)].value == DATE):
                        break;

                accounts_num = str(i)
                print accounts_num

                sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

                sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
                sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
                sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
                sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
                sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
                sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
                sheet_assessReport['Y' + l] = 0
                sheet_assessReport['Z' + l] = 0
                sheet_assessReport['AA' + l] = 0
                sheet_assessReport['AB' + l] = 0
                sheet_assessReport['AC' + l] = 0
                sheet_assessReport['AD' + l] = 0
                sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
                sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
                sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

                sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

                sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
                sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
                sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
                sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
                sheet_assessReport['AM' + l] = 0
                sheet_assessReport['AN' + l] = 0
                sheet_assessReport['AO' + l] = 0
                sheet_assessReport['AP' + l] = 0
                sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

                last_date = key
                l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_4(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    f_list = []
    for item in file_list:
        pattern = re.compile(ur'证券投资基金估值表_山海泽湶4号私募投资基金_.*')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
    d_list = []
    print f_list
    for item in f_list:
        d = item.split('_')[-1].split('.')[0]
        d_list = d_list + [int(d.replace('-',''))]
    print d_list
    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)


    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    print last_date
    for key, item in date_dict:
        print key
        print item
        if key in [20160731]:
            continue
        if (key > last_date) and (key <= int(date.replace('/', ''))):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name('Sheet1')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'期货清算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
            if u'存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['E' + l] = 0
            index = sheet1.col_values(0).index(u'证券投资合计:')
            sheet_assessReport['F' + l] = sheet1.cell(index, 7).value

            index = sheet1.col_values(0).index(u'其中股票投资:')
            sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
            index = sheet1.col_values(0).index(u'其中基金投资:')
            sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金账户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
            if u'应付估值服务费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付估值服务费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(0).index(u'今日单位净值：')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'基金资产净值:' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'基金资产净值:')
                sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_5(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    dir_list = list(os.walk(assessreport_path))[0][1]
    f_list = []
    for item in file_list:
        pattern = re.compile(ur'SM0117_星辰之鸣石多策略五号证券投资基金.*')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
    print f_list
    d_list = []
    for item in f_list:
        d = item.split('.')[0][-8::]
        try:
            int(d)
        except:
            continue
        d_list = d_list + [int(d)]

    print f_list
    print d_list

    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)


    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if key < 20161219:
            if (key > last_date) and (key <= int(date.replace('/', ''))):
                assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
                sheet1 = assessReport_wb.sheet_by_name('Sheet1')
                DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
                sheet_assessReport['A' + l] = DATE
                if u'银行存款' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'银行存款')
                    sheet_assessReport['B' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['B' + l] = 0
                sheet_assessReport['C' + l] = '=D' + l + '+E' + l
                index = sheet1.col_values(2).index(u'清算备付金')
                sheet_assessReport['D' + l] = sheet1.cell(index, 8).value
                if u'存出保证金' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'存出保证金')
                    sheet_assessReport['E' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['E' + l] = 0
                index = sheet1.col_values(1).index(u'证券投资合计:')
                sheet_assessReport['F' + l] = sheet1.cell(index, 8).value

                index = sheet1.col_values(1).index(u'其中股票投资:')
                sheet_assessReport['G' + l] = sheet1.cell(index, 8).value
                index = sheet1.col_values(1).index(u'其中基金投资:')
                sheet_assessReport['H' + l] = sheet1.cell(index, 8).value
                if u'买入返售金额资产' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'买入返售金额资产')
                    sheet_assessReport['I' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['I' + l] = 0

                if u'证券资金账户' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'证券资金账户')
                    sheet_assessReport['J' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['J' + l] = 0

                if u'理财产品' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'理财产品')
                    sheet_assessReport['K' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['K' + l] = 0

                index = sheet1.col_values(2).index(u'应付托管费')
                sheet_assessReport['L' + l] = sheet1.cell(index, 8).value
                if u'应付基金运营服务费' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'应付基金运营服务费')
                    sheet_assessReport['M' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['M' + l] = 0
                if u'应付管理人报酬' in sheet1.col_values(2):
                    index = sheet1.col_values(2).index(u'应付管理人报酬')
                    sheet_assessReport['N' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['N' + l] = 0
                index = sheet1.col_values(1).index(u'今日单位净值：')
                sheet_assessReport['O' + l] = sheet1.cell(index, 8).value

                if u'基金资产净值:' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'基金资产净值:')
                    sheet_assessReport['P' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['P' + l] = 0

                index = sheet1.col_values(1).index(u'实收资本')
                sheet_assessReport['Q' + l] = sheet1.cell(index, 8).value

                print DATE
                i = 2;
                while 1:
                    i = i + 1
                    if (sheet_accounts["A" + str(i)].value == DATE):
                        break;

                accounts_num = str(i)
                print accounts_num

                sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

                sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
                sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
                sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
                sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
                sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
                sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
                sheet_assessReport['Y' + l] = 0
                sheet_assessReport['Z' + l] = 0
                sheet_assessReport['AA' + l] = 0
                sheet_assessReport['AB' + l] = 0
                sheet_assessReport['AC' + l] = 0
                sheet_assessReport['AD' + l] = 0
                sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
                sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
                sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

                sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

                sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
                sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
                sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
                sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
                sheet_assessReport['AM' + l] = 0
                sheet_assessReport['AN' + l] = 0
                sheet_assessReport['AO' + l] = 0
                sheet_assessReport['AP' + l] = 0
                sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

                last_date = key
                l = str(int(l) + 1)
        else:
            if (key > last_date) and (key <= int(date.replace('/', ''))):
                assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
                sheet1 = assessReport_wb.sheet_by_name('Sheet1')
                DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
                sheet_assessReport['A' + l] = DATE
                if u'银行存款' in sheet1.col_values(0):
                    index = sheet1.col_values(0).index(u'银行存款')
                    sheet_assessReport['B' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['B' + l] = 0
                sheet_assessReport['C' + l] = '=D' + l + '+E' + l
                index = sheet1.col_values(1).index(u'清算备付金')
                sheet_assessReport['D' + l] = sheet1.cell(index, 7).value
                if u'存出保证金' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'存出保证金')
                    sheet_assessReport['E' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['E' + l] = 0
                index = sheet1.col_values(0).index(u'证券投资合计:')
                sheet_assessReport['F' + l] = sheet1.cell(index, 7).value

                index = sheet1.col_values(0).index(u'其中股票投资:')
                sheet_assessReport['G' + l] = sheet1.cell(index, 7).value
                index = sheet1.col_values(0).index(u'其中基金投资:')
                sheet_assessReport['H' + l] = sheet1.cell(index, 7).value
                if u'买入返售金额资产' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'买入返售金额资产')
                    sheet_assessReport['I' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['I' + l] = 0

                if u'证券资金账户' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'证券资金账户')
                    sheet_assessReport['J' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['J' + l] = 0

                if u'理财产品' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'理财产品')
                    sheet_assessReport['K' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['K' + l] = 0

                index = sheet1.col_values(1).index(u'应付托管费')
                sheet_assessReport['L' + l] = sheet1.cell(index, 7).value
                if u'应付基金运营服务费' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'应付基金运营服务费')
                    sheet_assessReport['M' + l] = sheet1.cell(index, 8).value
                else:
                    sheet_assessReport['M' + l] = 0
                if u'应付管理人报酬' in sheet1.col_values(1):
                    index = sheet1.col_values(1).index(u'应付管理人报酬')
                    sheet_assessReport['N' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['N' + l] = 0
                index = sheet1.col_values(0).index(u'今日单位净值：')
                sheet_assessReport['O' + l] = sheet1.cell(index, 7).value

                if u'基金资产净值:' in sheet1.col_values(0):
                    index = sheet1.col_values(0).index(u'基金资产净值:')
                    sheet_assessReport['P' + l] = sheet1.cell(index, 7).value
                else:
                    sheet_assessReport['P' + l] = 0

                index = sheet1.col_values(0).index(u'实收资本')
                sheet_assessReport['Q' + l] = sheet1.cell(index, 7).value

                print DATE
                i = 2;
                while 1:
                    i = i + 1
                    if (sheet_accounts["A" + str(i)].value == DATE):
                        break;

                accounts_num = str(i)
                print accounts_num

                sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

                sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
                sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
                sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
                sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
                sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
                sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
                sheet_assessReport['Y' + l] = 0
                sheet_assessReport['Z' + l] = 0
                sheet_assessReport['AA' + l] = 0
                sheet_assessReport['AB' + l] = 0
                sheet_assessReport['AC' + l] = 0
                sheet_assessReport['AD' + l] = 0
                sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
                sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
                sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

                sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

                sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
                sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
                sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
                sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
                sheet_assessReport['AM' + l] = 0
                sheet_assessReport['AN' + l] = 0
                sheet_assessReport['AO' + l] = 0
                sheet_assessReport['AP' + l] = 0
                sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

                last_date = key
                l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_7(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    dir_list = list(os.walk(assessreport_path))[0][1]
    f_list = []
    for item in file_list:
        pattern = re.compile(ur'MSLH07鸣石量化七号私募投资基金委托资产资产估值表.*')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
    d_list = []
    for item in f_list:
        d = item.split('.')[0][-8::]
        try:
            int(d)
        except:
            continue
        d_list = d_list + [int(d)]

    print f_list
    print d_list

    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)


    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if (key > last_date) and (key <= int(date.replace('/', ''))):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name(u'第1页')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'结算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 11).value
            if u'存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['E' + l] = 0
            index = sheet1.col_values(0).index(u'证券投资合计')
            sheet_assessReport['F' + l] = sheet1.cell(index, 11).value

            index = sheet1.col_values(0).index(u'其中股票投资')
            sheet_assessReport['G' + l] = sheet1.cell(index, 11).value
            if u'其中基金投资' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'其中基金投资')
                sheet_assessReport['H' + l] = sheet1.cell(index, 11).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金账户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 11).value
            if u'应付外包服务费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付外包服务费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(0).index(u'今日单位净值')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'基金资产净值' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'基金资产净值')
                sheet_assessReport['P' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 11).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))

def update_assessReport_11(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    dir_list = list(os.walk(assessreport_path))[0][1]
    f_list = []
    for item in file_list:
        pattern = re.compile(ur'SM6672_鸣石量化十一号私募投资基金_产品估值表_日报_.*')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
    d_list = []
    for item in f_list:
        d = item.split('_')[-1].split('.')[0]
        d_list = d_list + [int(d.replace('-',''))]

    print f_list
    print d_list

    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)


    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if (key > last_date) and (key <= int(date.replace('/', ''))) and (key < 20170119):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name(u'第1页')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'结算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 11).value
            if u'存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['E' + l] = 0

            sheet_assessReport['F' + l] = 'G' + l + 'H' + l

            index = sheet1.col_values(0).index(u'流通股票投资合计')
            sheet_assessReport['G' + l] = sheet1.cell(index, 11).value
            if u'其中基金投资' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'其中基金投资')
                sheet_assessReport['H' + l] = sheet1.cell(index, 11).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金账户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 11).value
            if u'应付销售服务费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付销售服务费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(0).index(u'单位净值')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'基金资产净值' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'基金资产净值')
                sheet_assessReport['P' + l] = sheet1.cell(index, 11).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 11).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
        if (key > last_date) and (key <= int(date.replace('/', ''))) and (key >= 20170119):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name(u'第1页')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 6).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'结算备付金')
            sheet_assessReport['D' + l] = sheet1.cell(index, 6).value
            if u'存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 6).value
            else:
                sheet_assessReport['E' + l] = 0

            sheet_assessReport['F' + l] = 'G' + l + 'H' + l

            index = sheet1.col_values(0).index(u'流通股票投资合计')
            sheet_assessReport['G' + l] = sheet1.cell(index, 6).value
            if u'其中基金投资' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'其中基金投资')
                sheet_assessReport['H' + l] = sheet1.cell(index, 6).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 6).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金账户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 6).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 6).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 6).value
            if u'应付销售服务费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付销售服务费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 6).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 6).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(0).index(u'单位净值')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'基金资产净值' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'基金资产净值')
                sheet_assessReport['P' + l] = sheet1.cell(index, 6).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 6).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))


def update_assessReport_19(product_id, assessreport_path, date):
    file_list = list(os.walk(assessreport_path))[0][2]
    dir_list = list(os.walk(assessreport_path))[0][1]
    f_list = []
    for item in file_list:
        pattern = re.compile(ur'X2488S5鸣石量化十九号私募证券投资基金委托资产资产估值表.*')
        if len(pattern.findall(item)) != 0:
            f_list = f_list + [item]
    d_list = []
    for item in f_list:
        d = item.split('.')[0][-8::]
        try:
            int(d)
        except:
            continue
        d_list = d_list + [int(d)]

    print f_list
    print d_list

    date_dict = dict(zip(d_list, f_list))
    date_dict = sorted(date_dict.iteritems(), key = lambda x: x[0], reverse=False)


    wb = ox.load_workbook(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
    sheet_assessReport = wb["assessReport"]
    sheet_accounts = wb['accounts']
    i = 3
    while 1:
        i = i + 1
        if not sheet_assessReport["A" + str(i)].value:
            break
    l = str(i)
    last_date = sheet_assessReport["A" + str(i - 1)].value
    last_date = last_date.replace('/', '')
    last_date = int(last_date)
    for key, item in date_dict:
        print key
        print item
        if (key > last_date) and (key <= int(date.replace('/', ''))):
            assessReport_wb = xlrd.open_workbook(ur'{}/{}'.format(assessreport_path, item))
            sheet1 = assessReport_wb.sheet_by_name(u'第1页')
            DATE = str(key)[:4] + '/' + str(key)[4:6] + '/' + str(key)[6:8]
            sheet_assessReport['A' + l] = DATE
            if u'银行存款' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'银行存款')
                sheet_assessReport['B' + l] = sheet1.cell(index, 9).value
            else:
                sheet_assessReport['B' + l] = 0
            sheet_assessReport['C' + l] = '=D' + l + '+E' + l
            index = sheet1.col_values(1).index(u'结算备付金_期货备付金账户')
            sheet_assessReport['D' + l] = sheet1.cell(index, 9).value
            if u'存出保证金' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'存出保证金')
                sheet_assessReport['E' + l] = sheet1.cell(index, 9).value
            else:
                sheet_assessReport['E' + l] = 0

            sheet_assessReport['F' + l] = '=G' + l + '+H' + l

            index = sheet1.col_values(0).index(u'流通股票投资合计')
            sheet_assessReport['G' + l] = sheet1.cell(index, 9).value
            if u'其中基金投资' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'其中基金投资')
                sheet_assessReport['H' + l] = sheet1.cell(index, 9).value
            if u'买入返售金额资产' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'买入返售金额资产')
                sheet_assessReport['I' + l] = sheet1.cell(index, 9).value
            else:
                sheet_assessReport['I' + l] = 0

            if u'证券资金账户' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'证券资金账户')
                sheet_assessReport['J' + l] = sheet1.cell(index, 9).value
            else:
                sheet_assessReport['J' + l] = 0

            if u'理财产品' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'理财产品')
                sheet_assessReport['K' + l] = sheet1.cell(index, 9).value
            else:
                sheet_assessReport['K' + l] = 0

            index = sheet1.col_values(1).index(u'应付托管费')
            sheet_assessReport['L' + l] = sheet1.cell(index, 9).value
            if u'应付投资顾问费' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付投资顾问费')
                sheet_assessReport['M' + l] = sheet1.cell(index, 9).value
            else:
                sheet_assessReport['M' + l] = 0
            if u'应付管理人报酬' in sheet1.col_values(1):
                index = sheet1.col_values(1).index(u'应付管理人报酬')
                sheet_assessReport['N' + l] = sheet1.cell(index, 9).value
            else:
                sheet_assessReport['N' + l] = 0
            index = sheet1.col_values(0).index(u'今日单位净值')
            sheet_assessReport['O' + l] = sheet1.cell(index, 1).value

            if u'基金资产净值' in sheet1.col_values(0):
                index = sheet1.col_values(0).index(u'基金资产净值')
                sheet_assessReport['P' + l] = sheet1.cell(index, 9).value
            else:
                sheet_assessReport['P' + l] = 0

            index = sheet1.col_values(0).index(u'实收资本')
            sheet_assessReport['Q' + l] = sheet1.cell(index, 9).value

            print DATE
            i = 2;
            while 1:
                i = i + 1
                if (sheet_accounts["A" + str(i)].value == DATE):
                    break;

            accounts_num = str(i)
            print accounts_num

            sheet_assessReport['R' + l] = '=accounts!M' + accounts_num

            sheet_assessReport['S' + l] = '=accounts!L' + accounts_num
            sheet_assessReport['T' + l] = '=accounts!B' + accounts_num
            sheet_assessReport['U' + l] = '=accounts!G' + accounts_num
            sheet_assessReport['V' + l] = '=accounts!I' + accounts_num
            sheet_assessReport['W' + l] = '=accounts!H' + accounts_num
            sheet_assessReport['X' + l] = '=accounts!F' + accounts_num
            sheet_assessReport['Y' + l] = 0
            sheet_assessReport['Z' + l] = 0
            sheet_assessReport['AA' + l] = 0
            sheet_assessReport['AB' + l] = 0
            sheet_assessReport['AC' + l] = 0
            sheet_assessReport['AD' + l] = 0
            sheet_assessReport['AE' + l] = '=accounts!AL' + accounts_num
            sheet_assessReport['AF' + l] = '=accounts!AM' + accounts_num
            sheet_assessReport['AG' + l] = '=accounts!AK' + accounts_num

            sheet_assessReport['AH' + l] = '=accounts!AP' + accounts_num

            sheet_assessReport['AI' + l] = '=accounts!U' + accounts_num
            sheet_assessReport['AJ' + l] = '=F' + l + '-S' + l
            sheet_assessReport['AK' + l] = '=C' + l + '-R' + l
            sheet_assessReport['AL' + l] = '=AJ' + l + '+AK' + l
            sheet_assessReport['AM' + l] = 0
            sheet_assessReport['AN' + l] = 0
            sheet_assessReport['AO' + l] = 0
            sheet_assessReport['AP' + l] = 0
            sheet_assessReport['AQ' + l] = '=O' + l + '-AI' + l

            last_date = key
            l = str(int(l) + 1)
    wb.save(u'//SHIMING/accounts/python/YZJ/accounts/accounts-{}/{}/accounts.xlsx'.format(product_id, date))
